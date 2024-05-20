import xlsxwriter
from os import startfile
from os import path
import openpyxl as xl
from openpyxl import utils
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog as fd


def select_read_file():
    global template_filename
    filetypes = [("Excel files", ".xlsx")]
    template_filename_temp = fd.askopenfilename(
        title="Открыть шаблон",
        filetypes=filetypes)
    if template_filename_temp != "":
        template_filename = fr"{template_filename_temp}"
        lbl_open_file_name["text"] = shorten_filename(template_filename)


def select_save_file():
    global report_filename
    filetypes = [("Excel files", ".xlsx")]
    report_filename_temp = fd.asksaveasfilename(
        title="Сохранить отчет",
        filetypes=filetypes)
    if report_filename_temp != "":
        report_filename = fr"{report_filename_temp}"
        lbl_report_file_name["text"] = shorten_filename(report_filename)


def shorten_filename(filename):
    pos_slash1 = filename.rfind("/")
    pos_slash2 = filename.rfind("/", 0, pos_slash1)
    short_name = "..." + filename[pos_slash2:]
    return short_name


def str_cm(num_float):
    str_dot = str(num_float)
    cm = str_dot.replace(".", ",")
    if cm.endswith(",0"):
        cm = cm.replace(",0", "")
    return cm


def excel_address(row, column):
    return f"{utils.cell.get_column_letter(column)}{row}"


def get_rich_string_old(equation, form_sub, form_super, form_def, form_bold):
    if equation.find("|") == -1 and equation.find("^") == -1 and equation.find("?") == -1:
        return -1
    rich_string = []
    i = 0
    while len(equation) > 0:
        if equation.find("|") == -1 and equation.find("^") == -1 and equation.find("?") == -1:
            rich_string.append(form_def)
            rich_string.append(equation)
            rich_string.append(form_def)
            return rich_string

        if equation[i] == "|" or equation[i] == "^" or equation[i] == "?":
            pos_start = i
            symbol = equation[i]
            pos_stop = equation.find(symbol, pos_start + 1)
            alt_str = equation[pos_start + 1:pos_stop]
            equation = equation[:pos_start] + equation[pos_stop + 1:]
            if pos_start > 0:
                def_str = equation[:pos_start]
                equation = equation[pos_start:]
                rich_string.append(form_def)
                rich_string.append(def_str)
            if symbol == "|":
                rich_string.append(form_sub)
            elif symbol == "^":
                rich_string.append(form_super)
            elif symbol == "?":
                rich_string.append(form_bold)
            if symbol == "|":
                rich_string.append(alt_str.upper())
            else:
                rich_string.append(alt_str)
            i = 0
        else:
            i += 1
    rich_string.append(form_def)
    return rich_string


def get_rich_string(equation, form_def, form_sub, form_super, form_bold, form_sub_bold, form_super_bold, address):
    """Возвращает готовые переменные в виде 'строка 1', 'стиль 1' ... 'строка N', 'стиль N' для размеченной строки"""
    symbols = ['|', '?', '^']

    def style(s1, s2):  # Комбинирование стилей
        if s2 not in s1:
            return s1 + s2
        else:
            return s1.replace(s2, "")

    def style_to_form(symbol):  # Перевод стиля в форму
        try:
            styles = {"": form_def, "|": form_sub, "^": form_super,
                      "?": form_bold, "?|": form_sub_bold, "?^": form_super_bold,
                      "|?": form_sub_bold, "^?": form_super_bold}
            form = styles[symbol]
        except KeyError:
            messagebox.showerror(message=f"Проверь ячейку {excel_address(*address)}")
            return
        return form

    def len_interval(interval):  # Длина интервала
        return interval[1] - interval[0]

    if all(equation.find(symbol) == -1 for symbol in symbols):  # Проверка, есть ли символы разметки
        return

    # Формирование строки без символов и позиций символов
    string = ''
    i = 0
    k = 0
    positions = []
    flag_upper = 0  # Флаг: 0 - нет подстрочного символа; 1 - был один подстрочный символ; 2 - два подстрочных символа
    # Проходим по всей строке и записываем в string сам текст, а в positions позицию символа и сам символ
    while any(equation.find(symbol) != -1 for symbol in symbols):
        if equation[i] in symbols:
            k += i
            positions.append([k, equation[i]])
            if equation[i] == "|":
                flag_upper += 1  # Флаг выставляется, если предыдущий символ был | и текущий тоже |, иначе сбрасывается
            else:
                flag_upper = 0
            if flag_upper == 2:  # Если выставлен флаг записываем в string капсов
                string += equation[:i].upper()
                flag_upper = 0
            else:
                string += equation[:i]  # В строку записывается выражение ДО символа
            equation = equation[i + 1:]
            i = 0
        else:
            i += 1
    string += equation

    # Если символов не четное количество
    if len(positions) % 2 != 0:
        messagebox.showerror(message=f"Проверь ячейку {excel_address(*address)}")
        return

    # Проверка скобок для моей тупой головехи
    if list(string).count('(') != list(string).count(')'):
        messagebox.showerror(message=f"Проверь скобки в ячейке {excel_address(*address)}")

    # Формирование интервалов
    intervals = [[0, positions[0][0], ""]]
    for i in range(len(positions) - 1):
        prev = intervals[i]
        next_pos = positions[i + 1]
        interval = [positions[i][0], next_pos[0], style(prev[2], positions[i][1])]
        intervals.append(interval)
    if intervals[-1][1] != len(string):
        intervals.append([intervals[-1][1], len(string), ''])
    empty_intervals = [interval for interval in intervals if len_interval(interval) == 0]
    for empty in empty_intervals:
        intervals.remove(empty)

    # Перевод стилей в формы
    for interval in intervals:
        interval[2] = style_to_form(interval[2])

    # Если интервал всего один, то возвращается стиль
    if len(intervals) == 1:
        return intervals[0][2]

    rich_string = []
    for interval in intervals:
        rich_string.append(interval[2])
        rich_string.append(string[interval[0]: interval[1]])
    rich_string.append(form_def)

    return rich_string


def insert_values(string_row, variables, address):
    if string_row.find("{") == -1:
        return string_row
    while True:
        pos_start = string_row.find("{")
        if pos_start == -1:
            break
        pos_stop = string_row.find("}")
        var_name = string_row[pos_start + 1:pos_stop]
        try:
            string_row = string_row[:pos_start] + str_cm(variables[var_name]) + string_row[pos_stop + 1:]
        except KeyError:
            messagebox.showerror(message=f"Нет такого значения! {var_name}\nВ ячейке {excel_address(*address)}")
            return
    return string_row


def parce_file(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb.active

    boundaries = utils.cell.range_boundaries(sheet.dimensions)
    row_max = boundaries[-1] + 1
    column_max = boundaries[-2] + 1

    column_widths = [
        sheet.column_dimensions[utils.cell.get_column_letter(i)].width - 0.7109375
        for i in range(3, column_max)]
    merged_cells = [list(merged.bounds) for merged in sheet.merged_cells.ranges]
    merged_cells = [[merged[1] - 1, merged[0] - 3, merged[3] - 1, merged[2] - 3] for merged in merged_cells]

    text = []
    variables = {}
    rows = []
    columns = []
    duplicates = []
    for i in range(1, row_max):
        var_names_unsplit = str(sheet.cell(row=i, column=1).value)
        var_values_unsplit = str(sheet.cell(row=i, column=2).value)
        if var_names_unsplit != "None":
            var_names = var_names_unsplit.split("\n")
            var_values_unsplit = var_values_unsplit.replace(",", ".")
            var_values = var_values_unsplit.split("\n")
            if len(var_names) != len(var_values):
                messagebox.showerror(message="Не хватает значений!")
                return
            for var_name, var_value in zip(var_names, var_values):
                if var_name != "None" and var_value != "None":
                    if var_name in variables:
                        duplicates.append(var_name)
                    var = {var_name: var_value}
                    variables.update(var)

    for i in range(1, row_max):
        for j in range(3, column_max):
            text_cell = sheet.cell(row=i, column=j).value
            if text_cell is not None:
                text.append(str(text_cell))
                rows.append(i - 1)
                columns.append(j - 3)
            else:
                continue

    if len(duplicates) > 0:
        str_duplicates = ""
        for var_name in duplicates:
            str_duplicates += var_name + ", "
        str_duplicates = str_duplicates[:-2] + "."
        messagebox.showerror(message="Повторяются названия значений:\n" + str_duplicates)

    return text, variables, rows, columns, column_widths, column_max - 1, row_max - 1, merged_cells


def podstava():
    if report_filename.endswith(".xlsx") is False or template_filename.endswith(".xlsx") is False:
        messagebox.showerror(message="Проверь файлы!")
        return
    workbook = xlsxwriter.Workbook(report_filename)
    worksheet = workbook.add_worksheet()

    cursive_get = int(cursive.get())
    form_def = workbook.add_format(
        {"italic": cursive_get, "valign": "top", "font_name": "ISOCPEUR", "font_size": 12, "text_wrap": 1, "border": 1})
    form_sub = workbook.add_format(
        {"italic": cursive_get, "font_script": 2, "font_name": "ISOCPEUR", "font_size": 16, "text_wrap": 1,
         "border": 1})
    form_super = workbook.add_format(
        {"italic": cursive_get, "font_script": 1, "font_name": "ISOCPEUR", "font_size": 16, "text_wrap": 1,
         "border": 1})
    form_bold = workbook.add_format(
        {"italic": cursive_get, "valign": "top", "bold": 1, "font_name": "ISOCPEUR", "font_size": 12, "text_wrap": 1,
         "border": 1})
    form_sub_bold = workbook.add_format(
        {"italic": cursive_get, "font_script": 2, "bold": 1, "font_name": "ISOCPEUR", "font_size": 16, "text_wrap": 1,
         "border": 1})
    form_super_bold = workbook.add_format(
        {"italic": cursive_get, "font_script": 1, "bold": 1, "font_name": "ISOCPEUR", "font_size": 16, "text_wrap": 1,
         "border": 1})
    form_bundle = [form_def, form_sub, form_super, form_bold, form_sub_bold, form_super_bold]

    parced_file = parce_file(template_filename)
    if parced_file == None:
        return
    text, variables, rows, columns, column_widths, column_max, row_max, merged_cells = parced_file

    for i, width in enumerate(column_widths, 0):
        worksheet.set_column(i, i, width)

    for merged in merged_cells:
        worksheet.merge_range(*merged, "", form_def)

    for row, column, cell in zip(rows, columns, text):
        cell_values = insert_values(cell, variables, (row + 1, column + 3))
        if cell_values == None:
            return
        rich_string = get_rich_string(cell_values, *form_bundle, (row + 1, column + 3))
        if rich_string == None:
            worksheet.write(row, column, cell_values, form_def)
        elif type(rich_string) == xlsxwriter.format.Format:
            cell_values = cell_values[1:-1]
            worksheet.write(row, column, cell_values, rich_string)
        else:
            worksheet.write_rich_string(row, column, *rich_string)

    try:
        workbook.close()
    except xlsxwriter.exceptions.FileCreateError as error:
        messagebox.showerror(message="Закрой файл отчета!")
        return

    report_path = path.abspath(report_filename)
    startfile(report_path)


window = tk.Tk()
window.title("Подставлятор3001")
window.resizable(width=False, height=False)

template_filename = r"template.xlsx"
report_filename = r"report.xlsx"

# Чтение файла

lbl_open_file_name = tk.Label(master=window, height=3, width=20, text=template_filename,
                              anchor="w", wraplength=200, justify=tk.LEFT)
lbl_open_file_name.grid(row=0, column=1, sticky="w")

btn_read_from_file = tk.Button(
    master=window,
    text="Шаблон:",
    width=7, height=2,
    command=select_read_file
)
btn_read_from_file.grid(row=0, column=0, padx=2, sticky="w")

# Сохранение фала

lbl_report_file_name = tk.Label(master=window, height=3, width=20, text=report_filename,
                                anchor="w", wraplength=200, justify=tk.LEFT)
lbl_report_file_name.grid(row=1, column=1, sticky="w")

btn_save_to_file = tk.Button(
    master=window,
    text="Отчет:",
    width=7, height=2,
    command=select_save_file
)
btn_save_to_file.grid(row=1, column=0, padx=2, sticky="w")

# Пуск

btn_podstava = tk.Button(
    master=window,
    text="Подставить!",
    width=30, height=2,
    command=podstava
)
btn_podstava.grid(row=2, column=0, padx=2, pady=2, columnspan=2)

# Курсив

cursive = tk.IntVar()
chb_cursive = tk.Checkbutton(master=window, text="Курсив",
                             offvalue=0, onvalue=1, variable=cursive)
chb_cursive.grid(row=3, column=0, columnspan=2, padx=0, sticky="w")

tk.mainloop()
